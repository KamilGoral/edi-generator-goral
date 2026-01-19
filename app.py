# app.py
import io
import os
import re
import unicodedata
import datetime
import tempfile
import hashlib
from pathlib import Path

import pandas as pd
import streamlit as st

# Pandas: jawne rzutowania (przyszłe zachowanie)
pd.set_option("future.no_silent_downcasting", True)

# ======================
# Utils
# ======================

def convert_xls_bytes_to_xlsx_python(file_bytes: bytes) -> str:
    """
    Konwersja .xls do .xlsx używając czystego Pythona (xlrd + openpyxl).
    Działa na każdym systemie bez Microsoft Excel.
    Zwraca ścieżkę do .xlsx.
    """
    try:
        import xlrd
        from openpyxl import Workbook
    except ImportError as e:
        raise RuntimeError("Wymagane pakiety: xlrd, openpyxl. Zainstaluj: pip install xlrd openpyxl") from e

    tmp_dir = tempfile.mkdtemp(prefix="edi_xls2xlsx_")
    src_path = os.path.join(tmp_dir, "input.xls")
    dst_path = os.path.join(tmp_dir, "output.xlsx")

    # Zapisz bytes do pliku .xls
    with open(src_path, "wb") as f:
        f.write(file_bytes)

    # Otwórz plik .xls z xlrd
    try:
        xls_book = xlrd.open_workbook(src_path, formatting_info=False)
    except Exception as e:
        raise RuntimeError(f"Nie udało się otworzyć pliku .xls: {e}") from e

    # Utwórz nowy plik .xlsx z openpyxl
    xlsx_book = Workbook()
    xlsx_book.remove(xlsx_book.active)  # Usuń domyślny arkusz

    # Kopiuj wszystkie arkusze
    for sheet_index in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_index)
        xlsx_sheet = xlsx_book.create_sheet(title=xls_sheet.name)

        # Kopiuj dane wiersz po wierszu
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_idx, col_idx)
                cell_type = xls_sheet.cell_type(row_idx, col_idx)

                # Konwersja typów xlrd -> openpyxl
                if cell_type == xlrd.XL_CELL_DATE:
                    # Konwersja daty
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                        cell_value = datetime.datetime(*date_tuple)
                    except:
                        pass
                elif cell_type == xlrd.XL_CELL_EMPTY:
                    cell_value = None

                xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

    # Zapisz jako .xlsx
    try:
        xlsx_book.save(dst_path)
    except Exception as e:
        raise RuntimeError(f"Nie udało się zapisać pliku .xlsx: {e}") from e

    if not os.path.exists(dst_path):
        raise RuntimeError("Konwersja .xls → .xlsx nie powiodła się.")

    return dst_path

def strip_diacritics(s: str) -> str:
    if s is None:
        return ""
    return unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()

def norm_key(s: str) -> str:
    return strip_diacritics(s).strip().lower()

def fmt2(v: float) -> str:
    try:
        # Formatowanie z dwoma miejscami po przecinku + zamiana kropki na przecinek
        return f"{float(v):.2f}".replace(".", ",")
    except Exception:
        return "0,00"

def norm_ean(x) -> str:
    s = str(x).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s.replace(" ", "")

def parse_price(v):
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return None

def parse_vat(v, default_vat=23) -> int:
    """
    VAT musi być liczbą/decimalem. Jeśli puste/tekst — przyjmij domyślny.
    Obsługa 0.23 → 23.
    """
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return int(default_vat)
    try:
        x = float(str(v).replace(",", "."))
        if 0 < x < 1:
            x *= 100.0
        return int(round(x))
    except Exception:
        return int(default_vat)

# EAN: rozbij wiele kodów w jednej komórce na wiele rekordów
EAN_RE = re.compile(r"\d{8,}")  # EAN/GTIN >= 8 cyfr

def split_eans(cell) -> list[str]:
    if pd.isna(cell):
        return []
    return EAN_RE.findall(str(cell))

def ensure_unique_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Zapewnia unikalne nazwy kolumn, dopisując sufiksy .1, .2 itd."""
    if df.columns.is_unique:
        return df
    counts = {}
    new_cols = []
    for c in map(str, df.columns):
        if c not in counts:
            counts[c] = 0
            new_cols.append(c)
        else:
            counts[c] += 1
            new_cols.append(f"{c}.{counts[c]}")
    df = df.copy()
    df.columns = new_cols
    return df

# ======================
# Kolumny i aliasy
# ======================

ALIASES = {
    "produkt": ["produkt", "nazwa", "nazwa towaru"],
    "kod_ean": ["kod ean", "ean", "ean produktu", "kod", "gtin"],
    "cena": ["cena zakupu netto", "cena netto", "cena"],
    "dystrybutor": ["dystrybutor", "dystrybutorzy", "dostawca", "dystrybucja", "dystrybucja:"],
    "vat": ["vat", "stawka vat"],  # opcjonalne
}

def find_cols(df: pd.DataFrame) -> dict:
    found = {k: None for k in ALIASES.keys()}
    for col in df.columns:
        ck = norm_key(col)
        for key, aliases in ALIASES.items():
            if found[key] is not None:
                continue
            for a in aliases:
                ak = norm_key(a)
                if ck == ak or ak in ck:
                    found[key] = col
                    break
    return found

SUP_SPLIT_RE = re.compile(r"[;,|/]+")

def split_suppliers(cell) -> list:
    if pd.isna(cell):
        return []
    toks = [t.strip() for t in SUP_SPLIT_RE.split(str(cell)) if t.strip()]
    seen = set()
    out = []
    for t in toks:
        k = t.lower()
        if k not in seen:
            seen.add(k)
            out.append(t)
    return out

def extract_distinct_suppliers(series: pd.Series) -> list:
    seen = {}
    for v in series.dropna().tolist():
        for s in split_suppliers(v):
            k = s.lower()
            if k not in seen:
                seen[k] = s
    return sorted(list(seen.values()), key=lambda x: x.lower())

def mask_rows_for_supplier(df: pd.DataFrame, supplier_col: str, supplier_name: str) -> pd.Series:
    name_l = supplier_name.lower()
    return df[supplier_col].apply(
        lambda v: any(s.lower() == name_l for s in split_suppliers(v)) if pd.notna(v) else False
    )

def coerce_vat_column(df: pd.DataFrame, cols: dict, default_vat: int) -> pd.DataFrame:
    """Tworzy kolumnę _VAT_USED_ (int 23/8/5/0). Jeśli brak/tekst — default_vat."""
    if cols.get("vat"):
        df["_VAT_USED_"] = df[cols["vat"]].apply(lambda v: parse_vat(v, default_vat))
    else:
        df["_VAT_USED_"] = int(default_vat)
    return df

def df_to_view(df: pd.DataFrame) -> pd.DataFrame:
    """
    Kopia DataFrame bezpieczna do wyświetlenia w Streamlit/Arrow:
    - unikalne nagłówki
    - nagłówki jako string
    - wartości jako string[python]
    - NaN → ""
    """
    df = ensure_unique_columns(df)
    out = df.copy()
    out.columns = [str(c) for c in out.columns]
    out = out.astype("string[python]").fillna("")
    return out

# ======================
# EDI
# ======================

def edi_header(now: datetime.datetime, supplier_name: str, issuer: dict, lines_count: int) -> str:
    t = []
    t.append("TypPolskichLiter:LA")
    t.append("TypDok:CENTR_KONTRAKT_DOST")
    t.append(f"NrDok:KTRD/{now.strftime('%Y%m')}/{now.strftime('%H%M%S')}")
    t.append(f"Data:{now.strftime('%d.%m.%Y')}")
    t.append("DotyczyDok:AUTOGENEROWANY")
    t.append("Magazyn:Mag nr 1")
    t.append("SposobPlatn:GOT")
    t.append("TerminPlatn:0")
    t.append("IndeksCentralny:NIE")
    t.append(f"NazwaWystawcy:{issuer['name']}")
    t.append(f"AdresWystawcy:{issuer['street']}, {issuer['zip']} {issuer['city']}")
    t.append(f"KodWystawcy:{issuer['zip']}")
    t.append(f"MiastoWystawcy:{issuer['city']}")
    t.append(f"UlicaWystawcy:{issuer['street']}")
    t.append(f"KodKrajuWystawcy:{issuer['country']}")
    t.append(f"NIPWystawcy:{issuer['vat']}")
    t.append(f"BankWystawcy:{issuer['bank']}")
    t.append(f"KontoWystawcy:{issuer['account']}")
    t.append("NrWystawcyWSieciSklepow:0")
    t.append("WystawcaToCentralaSieci:1")
    t.append(f"NazwaOdbiorcy:{supplier_name}")
    t.append("KodKrajuOdbiorcy:PL")
    t.append("NIPOdbiorcy:")
    t.append("NrOdbiorcyWSieciSklepow:0")
    t.append("OdbiorcaToCentralaSieci:0")
    t.append(f"IloscLinii:{lines_count}")
    return "\r\n".join(t) + "\r\n"

def edi_footer(sum23: float, sum8: float, sum5: float) -> str:
    r = []
    r.append(f"Stawka:Vat{{23}}SumaNet{{{fmt2(sum23)}}}SumaVat{{{fmt2(sum23*0.23)}}}")
    r.append(f"Stawka:Vat{{8}}SumaNet{{{fmt2(sum8)}}}SumaVat{{{fmt2(sum8*0.08)}}}")
    r.append(f"Stawka:Vat{{5}}SumaNet{{{fmt2(sum5)}}}SumaVat{{{fmt2(sum5*0.05)}}}")
    r.append(f"DoZaplaty:{fmt2(sum23*1.23 + sum8*1.08 + sum5*1.05)}")
    return "\r\n".join(r)

# ======================
# Sanitizacja tekstu do EDI
# ======================

def sanitize_text_for_edi(s: str) -> str:
    """
    Czyści tekst przed wstawieniem do EDI:
    - usuwa/tnie po miękkich enterach / nowych liniach
    - zamienia problematyczne znaki (&, /, {, }, |, ;) na bezpieczniejsze
    - scala wielokrotne spacje
    """
    if s is None:
        return ""

    s = str(s)

    # Ujednolicenie końców linii
    s = s.replace("\r\n", "\n").replace("\r", "\n")

    # Bierzemy tylko pierwszą linię (często "Limit max" itp. jest w drugiej)
    s = s.split("\n")[0]

    # Zamiana problematycznych znaków
    # &  -> " i "
    # /  -> "-"
    # { } używane są w formacie EDI, więc usuwamy z treści
    # |  -> "-"
    # ;  -> "," (żeby nie mieszać z separatorami)
    replacements = {
        "&": " i ",
        "/": "-",
        "{": "(",
        "}": ")",
        "|": "-",
        ";": ",",
    }
    for bad, good in replacements.items():
        s = s.replace(bad, good)

    # Zbij wielokrotne białe znaki do jednej spacji
    import re
    s = re.sub(r"\s+", " ", s)

    return s.strip()


def build_edi_for_supplier(df: pd.DataFrame, cols: dict, supplier_name: str, default_vat: int, issuer: dict) -> str:
    now = datetime.datetime.now()
    rows = []
    for _, row in df.iterrows():
        suppliers = split_suppliers(row[cols["dystrybutor"]])
        if not any(s.lower() == supplier_name.lower() for s in suppliers):
            continue
        price = parse_price(row[cols["cena"]])
        if price is None or price <= 0:
            continue
        rows.append((row, price))

    sum23 = sum8 = sum5 = 0.0
    body_lines = []

    for row, price in rows:
        vat_val = int(row["_VAT_USED_"]) if "_VAT_USED_" in row else (
            parse_vat(row[cols["vat"]], default_vat) if cols.get("vat") else int(default_vat)
        )

        if vat_val == 23:
            sum23 += float(price)
        elif vat_val == 8:
            sum8 += float(price)
        elif vat_val == 5:
            sum5 += float(price)

        produkt_raw = str(row[cols["produkt"]]) if cols["produkt"] else ""
        produkt = sanitize_text_for_edi(produkt_raw)
        ean = norm_ean(row[cols["kod_ean"]]) if cols["kod_ean"] else ""


        line = (
            "Linia:"
            f"Nazwa{{{produkt}}}"
            f"Kod{{{ean}}}"
            f"Vat{{{vat_val}}}"
            "Jm{}"
            "Asortyment{}"
            "Sww{}"
            "PKWiU{}"
            "Ilosc{0}"
            f"Cena{{n{fmt2(price)}}}"
            "Wartosc{n0.00}"
            "IleWOpak{}"
            "IleKgL{}"
            "CenaSp{n0.00}"
        )
        body_lines.append(line)

    header = edi_header(now, supplier_name, issuer, lines_count=len(body_lines))
    body = ("\r\n".join(body_lines) + "\r\n") if body_lines else ""
    footer = edi_footer(sum23, sum8, sum5)
    return header + body + footer

def sanitize_filename(s: str) -> str:
    base = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in s)
    return base or "supplier"

# ======================
# UI
# ======================

st.set_page_config(page_title="Generator EDI", page_icon="🧾", layout="wide")

# Fullscreen overlay holder + CSS
_overlay_holder = st.empty()
st.markdown("""
<style>
#_global_overlay_ {
  position: fixed;
  inset: 0;
  background: rgba(255,255,255,0.65);
  z-index: 9999;
  display: flex;
  align-items: center;
  justify-content: center;
  backdrop-filter: blur(2px);
  pointer-events: all; /* blokuje kliki pod spodem */
}
._overlay_box_ {
  padding: 18px 22px;
  border-radius: 10px;
  background: white;
  box-shadow: 0 6px 20px rgba(0,0,0,.12);
  font-weight: 600;
  font-size: 16px;
}
</style>
""", unsafe_allow_html=True)

def show_overlay(text="Przetwarzanie…"):
    _overlay_holder.markdown(
        f"""<div id="_global_overlay_"><div class="_overlay_box_">{text}</div></div>""",
        unsafe_allow_html=True
    )

def hide_overlay():
    _overlay_holder.empty()

# Uploader po polsku (CSS)
st.markdown("""
<style>
[data-testid="stFileUploaderDropzone"] div:first-child { display: none; }
[data-testid="stFileUploaderDropzone"]::before {
  content: "Przeciągnij i upuść plik lub kliknij „Przeglądaj pliki”";
  display: block; text-align: center; padding: .5rem 0; font-weight: 600;
}
[data-testid="stFileUploaderBrowseFiles"] span { visibility: hidden; }
[data-testid="stFileUploaderBrowseFiles"] span::after {
  content: "Przeglądaj pliki"; visibility: visible; position: absolute; left: 0; right: 0;
}
</style>
""", unsafe_allow_html=True)

with st.expander("Instrukcja", expanded=False):
    st.markdown("""
**Jak przerobić plik z Lewiatana:**
1. **Wgraj** plik (XLS/XLSX/ODS/CSV). Dla **XLS** konwersja do **XLSX** wykona się automatycznie przez **Microsoft Excel**.
2. Wybierz **arkusz** (jeśli jest ich kilka).
3. Wskaż **wiersz nagłówków** (licząc od 1). System **podświetli** kolumny: Produkt, EAN, Cena, Dystrybutor, VAT.
4. Przejdź do **zakładek dystrybutorów** i pobierz **EDI (.txt)** (przycisk po prawej).
""")

with st.expander("Ustawienia", expanded=False):
    st.caption("Kodowanie plików wyjściowych")
    st.text_input("Kodowanie", value="windows-1250", disabled=True, key="enc_ro")
    st.caption("Ustawienia wystawcy (opcjonalnie)")
    issuer_name = st.text_input("Nazwa wystawcy", "Góral i Wspólnicy Sp.k.", key="iss_name")
    issuer_city = st.text_input("Miasto", "ROCZYNY", key="iss_city")
    issuer_zip = st.text_input("Kod", "34-120", key="iss_zip")
    issuer_vat = st.text_input("NIP", "5512575479", key="iss_vat")
    issuer_street = st.text_input("Ulica", "BIELSKA 89", key="iss_street")
    issuer_country = st.text_input("Kraj (ISO-2)", "PL", key="iss_country")
    issuer_bank = st.text_input("Bank", "MBANK", key="iss_bank")
    issuer_account = st.text_input("Konto", "49114020040000340284745016", key="iss_account")
    default_vat = st.selectbox("Domyślny VAT (gdy brak kolumny VAT)", [23, 8, 5, 0], index=0, key="iss_defvat")

enc = "windows-1250"
issuer = {
    "name": issuer_name,
    "city": issuer_city,
    "zip": issuer_zip,
    "vat": issuer_vat,
    "street": issuer_street,
    "country": issuer_country,
    "bank": issuer_bank,
    "account": issuer_account,
}

# Session state: cache konwersji i identyfikacja pliku
if "uploaded_file_hash" not in st.session_state:
    st.session_state.uploaded_file_hash = None
if "converted_xlsx_path" not in st.session_state:
    st.session_state.converted_xlsx_path = None
if "base_in" not in st.session_state:
    st.session_state.base_in = None

# 1) Wgraj plik
f = st.file_uploader("Przeciągnij i upuść plik lub kliknij „Przeglądaj pliki”", type=["xls", "xlsx", "ods", "csv"])

df = None
selected_sheet = None
base_in = None

if f is not None:
    name = f.name
    lower = name.lower()
    file_bytes = f.getvalue()
    file_hash = hashlib.md5(file_bytes).hexdigest()

    # Nowy upload → reset cache konwersji
    if st.session_state.uploaded_file_hash != file_hash:
        st.session_state.uploaded_file_hash = file_hash
        st.session_state.converted_xlsx_path = None
        st.session_state.base_in = Path(name).stem

    base_in = st.session_state.base_in

    # 2) Jeśli .xls → konwertuj tylko raz, potem używaj cache
    if lower.endswith(".xls"):
        if st.session_state.converted_xlsx_path is None:
            try:
                with st.spinner("Konwertuję .xls → .xlsx…"):
                    show_overlay("Konwertuję plik .xls…")
                    xlsx_path = convert_xls_bytes_to_xlsx_python(file_bytes)
                    hide_overlay()
                st.session_state.converted_xlsx_path = xlsx_path
            except Exception as e:
                hide_overlay()
                st.error(f"Konwersja .xls → .xlsx nie powiodła się: {e}")
                st.stop()

        # Wybór arkusza z przekonwertowanego pliku
        xf = pd.ExcelFile(st.session_state.converted_xlsx_path, engine="openpyxl")
        selected_sheet = st.selectbox("Wybierz arkusz", xf.sheet_names, index=0) if len(xf.sheet_names) > 1 else xf.sheet_names[0]

    # 2b) Dla .xlsx/.ods — wybór arkusza normalnie
    elif lower.endswith((".xlsx", ".ods")):
        try:
            xf = pd.ExcelFile(io.BytesIO(file_bytes), engine=None)
            selected_sheet = st.selectbox("Wybierz arkusz", xf.sheet_names, index=0) if len(xf.sheet_names) > 1 else xf.sheet_names[0]
        except Exception as e:
            st.error(f"Nie udało się odczytać listy arkuszy: {e}")
            st.stop()
    else:
        selected_sheet = None  # CSV

    # 3) Wiersz nagłówków
    header_row = st.number_input("Wiersz nagłówków (licząc od 1)", min_value=1, value=1, step=1)

    # 4) Wczytaj dane (bez ponownej konwersji)
    try:
        with st.spinner("Wczytuję dane…"):
            show_overlay("Wczytywanie danych…")

            if lower.endswith(".xls"):
                df = pd.read_excel(
                    st.session_state.converted_xlsx_path,
                    header=header_row - 1,
                    engine="openpyxl",
                    sheet_name=selected_sheet
                )

            elif lower.endswith(".xlsx"):
                df = pd.read_excel(
                    io.BytesIO(file_bytes),
                    header=header_row - 1,
                    engine="openpyxl",
                    sheet_name=selected_sheet
                )

            elif lower.endswith(".ods"):
                df = pd.read_excel(
                    io.BytesIO(file_bytes),
                    header=header_row - 1,
                    sheet_name=selected_sheet
                )

            else:  # CSV z próbą kodowań
                last_err = None
                bio = io.BytesIO(file_bytes)
                for e_try in ("utf-8", "utf-8-sig", "cp1250", "latin1"):
                    try:
                        df = pd.read_csv(bio, header=header_row - 1, sep=None, engine="python", encoding=e_try)
                        last_err = None
                        break
                    except Exception as er:
                        last_err = er
                        bio.seek(0)
                if last_err is not None:
                    raise last_err

            hide_overlay()

        # Zapewnij unikalne nagłówki tuż po wczytaniu
        df = ensure_unique_columns(df)

    except Exception as e:
        hide_overlay()
        st.error(f"Błąd wczytywania: {e}")
        st.stop()

# 5) Podgląd + EDI
if df is not None:
    # Puste zamiast None/NaN (dla dalszych operacji)
    df = df.replace({None: pd.NA}).fillna(pd.NA)
    # Nagłówki jako string
    df.columns = [str(c) for c in df.columns]

    # Mapowanie kolumn
    cols = find_cols(df)

    # Rozbij wielo-EAN-owe wiersze
    if cols.get("kod_ean"):
        ean_col = cols["kod_ean"]
        df["_EANS_LIST_"] = df[ean_col].apply(lambda v: split_eans(v) if str(v).strip() else [])
        df["_EANS_LIST_"] = df.apply(
            lambda r: r["_EANS_LIST_"] if r["_EANS_LIST_"] else ([str(r[ean_col])] if str(r[ean_col]).strip() else []),
            axis=1
        )
        df = df.explode("_EANS_LIST_", ignore_index=True)
        df[ean_col] = df["_EANS_LIST_"].fillna("")
        df.drop(columns=["_EANS_LIST_"], inplace=True)

    # VAT do kolumny liczbowej
    df = coerce_vat_column(df, cols, int(default_vat))
    # Dla świętego spokoju jeszcze raz upewnij się co do unikalności nagłówków
    df = ensure_unique_columns(df)

    # Podgląd danych (z podświetleniem)
    st.write("Podgląd danych:")
    df_for_view = df.rename(columns={"_VAT_USED_": "VAT"})
    df_for_view = df_to_view(df_for_view)
    df_for_view = ensure_unique_columns(df_for_view)  # defensywnie przed Stylerem

    highlight_cols = [c for c in [
        cols.get("produkt"),
        cols.get("kod_ean"),
        cols.get("cena"),
        cols.get("dystrybutor"),
    ] if c] + (["VAT"] if "VAT" in df_for_view.columns else [])

    # Stylowanie tylko na istniejących kolumnach
    subset_cols = [c for c in highlight_cols if c in df_for_view.columns]
    styled = df_for_view.style.set_properties(
        subset=subset_cols,
        **{"background-color": "#FFF3CD"}
    )
    st.dataframe(styled, width="stretch")

    # Walidacja wymaganych kolumn
    missing = [k for k in ("produkt", "kod_ean", "cena", "dystrybutor") if cols.get(k) is None]
    if missing:
        st.error(f"Brakuje kolumn: {', '.join(missing)}. Upewnij się, że nagłówki odpowiadają aliasom.")
        st.stop()

    # Dystrybutorzy
    suppliers = extract_distinct_suppliers(df[cols["dystrybutor"]])
    if not suppliers:
        st.warning("Nie wykryto żadnych dystrybutorów w kolumnie DYSTRYBUTOR.")
        st.stop()

    # Liczba pozycji per dystrybutor (po cenie > 0) dla etykiety zakładki
    counts = []
    for sup in suppliers:
        m = mask_rows_for_supplier(df, cols["dystrybutor"], sup)
        dfx = df.loc[m].copy()
        if not dfx.empty:
            dfx["_parsed_price_"] = dfx[cols["cena"]].apply(parse_price)
            dfx = dfx[dfx["_parsed_price_"].apply(lambda x: x is not None and x > 0)]
        counts.append(len(dfx))

    tab_labels = [f"{sup} ({cnt})" for sup, cnt in zip(suppliers, counts)]
    tabs = st.tabs(tab_labels)

    base_in = base_in or "plik"

    for sup, tab in zip(suppliers, tabs):
        with tab:
            # Filtrowanie
            m = mask_rows_for_supplier(df, cols["dystrybutor"], sup)
            df_sup = df.loc[m].copy()

            # Podgląd (cena > 0)
            if not df_sup.empty:
                df_sup["_CenaParsed_"] = df_sup[cols["cena"]].apply(parse_price)
                df_sup = df_sup[df_sup["_CenaParsed_"].apply(lambda x: x is not None and x > 0)]

            # Generowanie EDI (na pełnym df, żeby sumy były spójne)
            edi_text = build_edi_for_supplier(df, cols, sup, default_vat=int(default_vat), issuer=issuer)
            edi_bytes = edi_text.encode("windows-1250", errors="replace")
            fname = f"{sanitize_filename(sup)}-{base_in}.txt"

            # Pasek tytuł + pobieranie
            t_left, t_right = st.columns([0.75, 0.25])
            with t_left:
                st.markdown(f"### Dystrybutor: {sup}")
            with t_right:
                st.download_button(
                    "⬇️ Pobierz EDI (.txt)",
                    data=edi_bytes,
                    file_name=fname,
                    mime="text/plain",
                    key=f"dl_{sanitize_filename(sup)}",
                    width="stretch"
                )

            # Podgląd produktów bez kolumny dystrybutora
            if df_sup.empty:
                st.info("Wszystkie pozycje mają cenę pustą lub ≤ 0 albo brak pozycji dla tego dystrybutora.")
            else:
                cols_to_show = [c for c in [
                    cols.get("produkt"),
                    cols.get("kod_ean"),
                    cols.get("cena"),
                ] if c] + ["_VAT_USED_"]
                pretty = df_sup[cols_to_show].rename(columns={
                    cols.get("produkt"): "Produkt",
                    cols.get("kod_ean"): "EAN",
                    cols.get("cena"): "Cena (netto)",
                    "_VAT_USED_": "VAT",
                })
                pretty = df_to_view(pretty)
                st.dataframe(pretty, width="stretch")